import csv
import os
from datetime import datetime
from tkinter import filedialog, messagebox


def _timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_txt(result: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        f.write(result.get("lead_template", "") + "\n\n")
        f.write("--- TRANSCRIPT ---\n")
        f.write(result.get("transcript", ""))


def export_csv(result: dict, path: str):
    checklist = result.get("checklist", [])
    rows = []
    for item in checklist:
        rows.append({
            "client": result.get("client", ""),
            "caller": result.get("caller_name", ""),
            "date": result.get("call_date", ""),
            "score": result.get("score", ""),
            "temp": result.get("preliminary_temp", ""),
            "checklist_item": item.get("item", ""),
            "result": item.get("result", ""),
            "note": item.get("note", ""),
        })
    if not rows:
        rows = [{"client": result.get("client", ""), "score": result.get("score", "")}]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def export_pdf(result: dict, path: str):
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Call Analysis Report", ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, f"Client: {result.get('client', '')}", ln=True)
    pdf.cell(0, 7, f"Caller: {result.get('caller_name', '')}   Date: {result.get('call_date', '')}", ln=True)
    pdf.cell(0, 7, f"Score: {result.get('score', 'N/A')} / 100   Temp: {result.get('preliminary_temp', 'N/A')}", ln=True)
    pdf.ln(4)

    # Lead template
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Lead Template", ln=True)
    pdf.set_font("Courier", "", 9)
    template = result.get("lead_template", "")
    for line in template.splitlines():
        pdf.cell(0, 5, line[:110], ln=True)
    pdf.ln(4)

    # Checklist
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Checklist", ln=True)
    pdf.set_font("Helvetica", "", 10)
    ICONS = {"yes": "[Y]", "no": "[N]", "partial": "[~]", "n/a": "[-]"}
    for item in result.get("checklist", []):
        icon = ICONS.get(item.get("result", "n/a").lower(), "[?]")
        text = f"{icon} {item.get('item', '')}"
        pdf.multi_cell(0, 6, text[:120])

    # Coaching
    coaching = result.get("coaching_notes", [])
    if coaching:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, "Coaching Notes", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for note in coaching:
            pdf.multi_cell(0, 6, f"• {note}")

    pdf.output(path)


def export_excel(result: dict, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="FFFFFF", bold=True)

    def hrow(label, value):
        row = ws.max_row + 1
        ws.cell(row, 1, label).font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, value)

    hrow("Client", result.get("client", ""))
    hrow("Caller", result.get("caller_name", ""))
    hrow("Date", result.get("call_date", ""))
    hrow("Score", result.get("score", ""))
    hrow("Temperature", result.get("preliminary_temp", ""))
    hrow("File", result.get("file", ""))

    # Sheet 2 — Checklist
    ws2 = wb.create_sheet("Checklist")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 30
    for item in result.get("checklist", []):
        ws2.append([item.get("result", ""), item.get("item", ""), item.get("note", "")])

    # Sheet 3 — Lead Template
    ws3 = wb.create_sheet("Lead Template")
    ws3.column_dimensions["A"].width = 80
    for line in result.get("lead_template", "").splitlines():
        ws3.append([line])

    # Sheet 4 — Transcript
    ws4 = wb.create_sheet("Transcript")
    ws4.column_dimensions["A"].width = 100
    for line in result.get("transcript", "").splitlines():
        ws4.append([line])

    wb.save(path)


# ── Dialog ────────────────────────────────────────────────────────────────────

def export_dialog(parent, result: dict):
    path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension=".txt",
        initialfile=f"call_{result.get('call_date', _timestamp()).replace('/', '-')}",
        filetypes=[
            ("Text file", "*.txt"),
            ("CSV file", "*.csv"),
            ("PDF file", "*.pdf"),
            ("Excel file", "*.xlsx"),
        ],
    )
    if not path:
        return

    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".csv":
            export_csv(result, path)
        elif ext == ".pdf":
            export_pdf(result, path)
        elif ext == ".xlsx":
            export_excel(result, path)
        else:
            export_txt(result, path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=parent)
    except Exception as e:
        messagebox.showerror("Export Error", str(e), parent=parent)
